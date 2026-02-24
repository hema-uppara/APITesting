import pandas as pd
from dotenv import load_dotenv
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font, PatternFill
from auto_mail import send_email

load_dotenv()

email_app_pswd = os.getenv('app_password')
sender_mail = os.getenv('sender_email')
recipient_mail = [os.getenv('recipient_mail'),]
sheet_name = os.getenv('sheet_name')
sheet_id = os.getenv('sheet_id')

current_datetime = datetime.now()
# Extract the month name and year
day = current_datetime.strftime('%d')
month_name = current_datetime.strftime('%B')  
year = current_datetime.strftime('%Y')  
output_file_name = f"House Report {day}-{month_name}-{year}"
current_date = current_datetime.strftime('%Y-%m-%d')

body = f"""
Dear StakeHolders,

Find attached the details for house prices in Nigeria as at {current_date}

Regards.
"""
subject = "HOUSE PRICE DETAILS for NIGERIA"

url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"

data = pd.read_csv(url)
house_details_df = {}

for title in data['house_title'].dropna().unique():
    # print(f'--- {title} ---\n')
    df = data.copy()
    house = df[df['house_title'] == title]
    house_details_df[title] = house.groupby(['house_title',
                                             'state']).agg({'Price':'mean'}).reset_index().sort_values(by=['house_title',
                                                                                                           'Price'], ascending=False)[:5].reset_index(drop=True)
    # print(house_details_df.items())

wb = Workbook()
ws1 = wb.active
ws1.title = 'HOUSE_PRICE_DETAILS'
ws1.merge_cells('B1:C1')
ws1['B1'] = 'Average House Prices in Nigeria'

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# header cell:
header_ranges = [('B1','C1')]
for cell_range in header_ranges:
    for row in ws1[cell_range[0]:cell_range[1]]:
        for cell in row:
            cell.fill = PatternFill(start_color='FFA9A9A9', end_color='FFA9A9A9', fill_type='solid')
            cell.font = Font(color="000000", bold=True)  # Set font color to black
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text
            cell.border = thin_border  # Apply border

ws1['A2'] = 'Detached Duplex'
ws1['A10'] = 'Semi Detached Duplex'
ws1['A18'] = 'Terraced Duplex'
ws1['A26'] = 'Detached Bungalow'
ws1['A34'] = 'House'
ws1['A42'] = 'Block of Flats'
ws1['A50'] = 'Terraced Bungalow'
ws1['A58'] = 'Semi-Detached Bungalow'

cells_to_bold = {
    'A2': 'Detached Duplex',
    'A10': 'Semi Detached Duplex',
    'A18': 'Terraced Duplex',
    'A26': 'Detached Bungalow',
    'A34': 'House',
    'A42': 'Block of Flats',
    'A50': 'Terraced Bungalow',
    'A58': 'Semi-Detached Bungalow'
}

# Set the cell values and make them bold
for cell, text in cells_to_bold.items():
    ws1[cell] = text
    ws1[cell].font = Font(bold=True)


dfs = [df.loc[:,['state','Price']] for df in house_details_df.values()]
start_row_sheets = [3,11,19,27,35,43,51,59]

for i, df_sheet1 in enumerate(dfs):
    # Write column names to Sheet1
    for col_idx, column_name in enumerate(df_sheet1.columns, start=2):
        cell = ws1.cell(row=start_row_sheets[i], column=col_idx, value=column_name)
        cell.fill = PatternFill(start_color='FF224677', end_color='FF224677', fill_type='solid')
        cell.font = Font(color="FFFFFF")  # Set font color to white
        cell.border = thin_border  # Apply border

    # Write data to Sheet1
    for r_idx, row in enumerate(df_sheet1.iterrows(), start=start_row_sheets[i] + 1):
        for c_idx, value in enumerate(row[1], start=2):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border  # Apply border

wb.save(f'{output_file_name}.xlsx')


send_email(sender_mail, email_app_pswd, recipient_mail, body, f'{output_file_name}', subject)